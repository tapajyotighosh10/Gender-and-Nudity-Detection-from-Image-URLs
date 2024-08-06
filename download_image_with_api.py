# import requests
# from openpyxl import Workbook
# from openpyxl.drawing.image import Image
# from PIL import Image as PILImage
# from io import BytesIO
# import os

# # List of image URLs (add your 466 URLs here)
# urls = [
#     "https://static1.mingle2.com/images/users/81/37/25949343_976.jpg?1529993526",
#     "https://static1.mingle2.com/images/users/44/23/48922400_4483.jpg?1604028895",
#     "https://cdn-v3.justsayhi.com/u/profile_photo/mOvbRf57eWj_large_optimized.jpeg",
#     "https://cdn-v3.justsayhi.com/u/profile_photo/A9Z342WUIHx_large_optimized.jpeg",
#     "https://cdn-v3.justsayhi.com/u/profile_photo/jOsHgy5KCC7_large_optimized.jpeg",
#     "https://cdn-v3.justsayhi.com/u/profile_photo/_qtRGmarvk6_large_optimized.jpeg"
# ]

# # Function to download an image and return it as a PIL image
# def download_image(url):
#     response = requests.get(url)
#     if response.status_code == 200:
#         return PILImage.open(BytesIO(response.content))
#     else:
#         return None

# # Function to analyze gender using the API
# def get_gender(image_path):
#     api_url = 'https://nucleuz-dev.swiftsecurity.ai/api/images/detect-gender'
#     files = {'file': open(image_path, 'rb')}
#     response = requests.post(api_url, files=files)
#     if response.status_code == 200:
#         result = response.json()
#         return result.get('value', 'Unknown')
#     return 'Unknown'

# # Function to analyze nudity using the API
# def get_nudity(image_path):
#     api_url = 'https://nucleuz-dev.swiftsecurity.ai/api/images/classify'
#     files = {'file': open(image_path, 'rb')}
#     response = requests.post(api_url, files=files)
#     if response.status_code == 200:
#         result = response.json()
#         return 'Yes' if result.get('nudity', False) else 'No'
#     return 'No'

# # Create a new Excel workbook and add a worksheet
# wb = Workbook()
# ws = wb.active
# ws.title = "Image Analysis"

# # Set column headers
# ws.append(["URL", "Image", "Filename", "Gender", "Race", "Nudity"])

# # Set column widths and row height
# ws.column_dimensions['A'].width = 60  # URL column width
# ws.column_dimensions['B'].width = 20  # Image column width
# ws.column_dimensions['C'].width = 20  # Filename column width
# ws.column_dimensions['D'].width = 15  # Gender column width
# ws.column_dimensions['E'].width = 15  # Race column width
# ws.column_dimensions['F'].width = 15  # Nudity column width

# # Set a fixed row height for images
# row_height = 100

# # Directory to save images
# image_dir = "./images/"
# os.makedirs(image_dir, exist_ok=True)

# # Iterate over the URLs
# for index, url in enumerate(urls):
#     img = download_image(url)
#     if img:
#         # Save the image locally and add it to the Excel sheet
#         img_filename = f"image_{index+1}.png"
#         img_path = os.path.join(image_dir, img_filename)
#         img.save(img_path)

#         # Get the image dimensions
#         image_width, image_height = img.size

#         # Resize image if needed (ensure it fits within the column width)
#         max_image_width = 200
#         max_image_height = row_height
#         if image_width > max_image_width or image_height > max_image_height:
#             img.thumbnail((max_image_width, max_image_height), PILImage.LANCZOS)
#             img_width, img_height = img.size
#         else:
#             img_width, img_height = image_width, image_height

#         # Save the resized image
#         img.save(img_path)

#         # Load the image into openpyxl
#         img_excel = Image(img_path)
#         img_excel.width = img_width  # Adjust the size as needed
#         img_excel.height = img_height

#         # Analyze the image for gender and nudity
#         gender = get_gender(img_path)
#         nudity = get_nudity(img_path)

#         # Append the URL, image, filename, and analysis to the worksheet
#         ws[f"A{index+2}"] = url
#         ws.add_image(img_excel, f"B{index+2}")
#         ws[f"C{index+2}"] = img_filename
#         ws[f"D{index+2}"] = gender
#         ws[f"E{index+2}"] = "Unknown"  # Replace with actual race detection if needed
#         ws[f"F{index+2}"] = nudity

#         # Set the row height to fit the image
#         ws.row_dimensions[index+2].height = row_height  # Fixed height for consistency
#     else:
#         ws.append([url, "Failed to download", "", "", "", ""])

# # Save the workbook
# wb.save("image_analysis.xlsx")


# import requests
# from openpyxl import Workbook
# from openpyxl.drawing.image import Image
# from PIL import Image as PILImage
# from io import BytesIO
# import os

# # List of image URLs (add your 466 URLs here)
# urls = [
#     "https://static1.mingle2.com/images/users/81/37/25949343_976.jpg?1529993526",
#     "https://static1.mingle2.com/images/users/44/23/48922400_4483.jpg?1604028895",
#     "https://cdn-v3.justsayhi.com/u/profile_photo/mOvbRf57eWj_large_optimized.jpeg",
#     "https://cdn-v3.justsayhi.com/u/profile_photo/A9Z342WUIHx_large_optimized.jpeg",
#     "https://cdn-v3.justsayhi.com/u/profile_photo/jOsHgy5KCC7_large_optimized.jpeg",
#     "https://cdn-v3.justsayhi.com/u/profile_photo/_qtRGmarvk6_large_optimized.jpeg"
#     # Add more URLs here
# ]

# # Function to download an image and return it as a PIL image
# def download_image(url):
#     response = requests.get(url)
#     if response.status_code == 200:
#         return PILImage.open(BytesIO(response.content))
#     else:
#         return None

# # Function to analyze image for gender
# def get_gender(img_path):
#     response = requests.post(
#         'https://nucleuz-dev.swiftsecurity.ai/api/images/detect-gender',
#         files={'file': open(img_path, 'rb')}
#     )
#     if response.status_code == 200:
#         result = response.json()
#         return result.get("value", "Unknown")
#     return "Unknown"

# # Function to analyze image for nudity
# def get_nudity(img_path):
#     response = requests.post(
#         'https://nucleuz-dev.swiftsecurity.ai/api/images/classify',
#         files={'file': open(img_path, 'rb')}
#     )
#     if response.status_code == 200:
#         result = response.json()
#         return "Yes" if result.get("nudity", False) else "No"
#     return "Unknown"

# # Create a new Excel workbook and add a worksheet
# wb = Workbook()
# ws = wb.active
# ws.title = "Image Analysis"

# # Set column headers
# ws.append(["URL", "Image", "Filename", "Gender", "Race", "Nudity"])

# # Set column widths
# ws.column_dimensions['A'].width = 75  # Adjust URL column width
# ws.column_dimensions['B'].width = 20  # Adjust image column width
# ws.column_dimensions['C'].width = 15  # Adjust Filename column width
# ws.column_dimensions['D'].width = 15  # Adjust Gender column width
# ws.column_dimensions['E'].width = 15  # Adjust Race column width
# ws.column_dimensions['F'].width = 15  # Adjust Nudity column width

# # Set a fixed row height for images
# row_height = 100

# # Create a directory to store images
# if not os.path.exists('images'):
#     os.makedirs('images')

# # Iterate over the URLs
# for index, url in enumerate(urls):
#     img = download_image(url)
#     if img:
#         # Save the image locally
#         img_filename = f"images/image_{index+1}.png"
#         img_path = img_filename
#         img.save(img_path)

#         # Get the image dimensions
#         image_width, image_height = img.size

#         # Resize image if needed (ensure it fits within the column width)
#         max_image_width = 100
#         max_image_height = 100
#         if image_width > max_image_width or image_height > max_image_height:
#             img.thumbnail((max_image_width, max_image_height), PILImage.LANCZOS)
#             img_width, img_height = img.size
#         else:
#             img_width, img_height = image_width, image_height

#         # Save the resized image
#         img.save(img_path)

#         # Load the image into openpyxl
#         img_excel = Image(img_path)
#         img_excel.width = img_width  # Adjust the size as needed
#         img_excel.height = img_height

#         # Get Gender and Nudity values
#         gender = get_gender(img_path)
#         nudity = get_nudity(img_path)

#         # Append the URL, image, filename, and analysis to the worksheet
#         ws[f"A{index+2}"] = url
#         ws.add_image(img_excel, f"B{index+2}")
#         ws[f"C{index+2}"] = img_filename
#         ws[f"D{index+2}"] = gender
#         ws[f"E{index+2}"] = "Unknown"  # Implement your own logic or API for race detection
#         ws[f"F{index+2}"] = nudity

#         # Set the row height to match the image height
#         ws.row_dimensions[index+2].height = row_height  # Fixed height for consistency
#     else:
#         ws.append([url, "Failed to download", "", "", "", ""])

# # Save the workbook
# wb.save("image_analysis.xlsx")


import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
from io import BytesIO
import os

urls = [
"https://static1.mingle2.com/images/users/81/37/25949343_976.jpg?1529993526",
 "https://static1.mingle2.com/images/users/44/23/48922400_4483.jpg?1604028895", 
 "https://cdn-v3.justsayhi.com/u/profile_photo/mOvbRf57eWj_large_optimized.jpeg",
  "https://cdn-v3.justsayhi.com/u/profile_photo/A9Z342WUIHx_large_optimized.jpeg",
   "https://cdn-v3.justsayhi.com/u/profile_photo/jOsHgy5KCC7_large_optimized.jpeg", 
   "https://cdn-v3.justsayhi.com/u/profile_photo/_qtRGmarvk6_large_optimized.jpeg",
    "https://cdn-v3.justsayhi.com/u/profile_photo/2CZTgmLhCQE_large_optimized.jpeg", 
    "https://cdn-v3.justsayhi.com/u/profile_photo/bB7Hs0oy_5L_large_optimized.jpeg", 
    "https://cdn-v3.justsayhi.com/u/profile_photo/-RwFSa0I_Id_large_optimized.jpeg", 
    "https://cdn-v3.justsayhi.com/u/profile_photo/b2OlzrAyrws_large_optimized.jpeg", 
    "https://cdn-v3.justsayhi.com/u/profile_photo/gY0zrXUSq7b_large_optimized.jpeg", 
    "https://cdn-v3.justsayhi.com/u/profile_photo/yPasa8mSo-f_large_optimized.jpeg", 
    "https://static1.mingle2.com/images/users/66/70/74072689_9245.jpg?1685582316", 
    "https://static1.mingle2.com/images/users/37/86/74112873_5744.jpg?1685757322",
     "https://static1.mingle2.com/images/users/14/10/74174565_6983.jpg?1694185008", 
     "https://static1.mingle2.com/images/users/6/73/74174565_69.jpg?1694185045", 
     "https://static1.mingle2.com/images/users/89/71/74174565_4718.jpg?1694185045",
      "https://static1.mingle2.com/images/users/95/39/74174565_9363.jpg?1694188008", 
      "https://static1.mingle2.com/images/users/88/97/74183136_9217.jpg?1692445475", 
      "https://static1.mingle2.com/images/users/67/39/74183136_9278.jpg?1692446200", 
       "https://static1.mingle2.com/images/users/73/3/74183136_5855.jpg?1696018755", 
       "https://static1.mingle2.com/images/users/22/37/74186567_6707.jpg?1686695327", 
       "https://static1.mingle2.com/images/users/46/98/74186567_6183.jpg?1686697056", 
       "https://static1.mingle2.com/images/users/65/16/74186567_7912.jpg?1686697056", 
       "https://static1.mingle2.com/images/users/52/62/74219211_2633.jpg?1686200063", 
       "https://cdn-v3.justsayhi.com/u/profile_photo/asA1EW8Hu_Y_large_optimized.jpeg", 
       "https://cdn-v3.justsayhi.com/u/profile_photo/jmBLUNJWH23_large_optimized.jpeg", 
       "https://static1.mingle2.com/images/users/89/62/74332103_5784.jpg?1690358500", 
       "https://static1.mingle2.com/images/users/73/88/74332103_5779.jpg?1690358538", 
       "https://static1.mingle2.com/images/users/89/24/74332103_2597.jpg?1691333012", 
       "https://static1.mingle2.com/images/users/37/33/74332103_9677.jpg?1691333012", 
       "https://static1.mingle2.com/images/users/45/91/74407136_4369.jpg?1686981954", 
       "https://static1.mingle2.com/images/users/4/47/74446088_479.jpg?1688065454", 
       "https://static1.mingle2.com/images/users/50/26/74446088_2676.jpg?1688065282", 
       "https://static1.mingle2.com/images/users/86/23/74523896_8993.jpg?1687481268", 
       "https://static1.mingle2.com/images/users/59/92/74524088_135.jpg?1687481798", 
       "https://static1.mingle2.com/images/users/99/0/74524275_5255.jpg?1687483183", 
       "https://static1.mingle2.com/images/users/45/42/74524275_7980.jpg?1687483356", 
       "https://static1.mingle2.com/images/users/15/77/74524405_788.jpg?1687483563", 
       "https://static1.mingle2.com/images/users/98/7/74524558_1440.jpg?1687484137", 
       "https://static1.mingle2.com/images/users/41/82/74583249_7558.jpg?1687749026",
        "https://static1.mingle2.com/images/users/15/67/74640797_4796.jpg?1688023184",
         "https://static1.mingle2.com/images/users/8/89/74640797_6983.jpg?1688028327", 
         "https://static1.mingle2.com/images/users/53/31/74640797_6515.jpg?1688028327",
          "https://static1.mingle2.com/images/users/13/88/74640797_731.jpg?1688028327",
           "https://static1.mingle2.com/images/users/10/93/74640797_4338.jpg?1688028327",
            "https://static1.mingle2.com/images/users/16/93/74640797_3255.jpg?1688028327",
             "https://static1.mingle2.com/images/users/3/67/74715894_1144.jpg?1688350793", 
             "https://static1.mingle2.com/images/users/17/34/74729326_6623.jpg?1688411619",
              "https://static1.mingle2.com/images/users/66/6/74784845_147.jpg?1688661742", 
              "https://static1.mingle2.com/images/users/39/76/74830572_2744.jpg?1688872160", 
              "https://static1.mingle2.com/images/users/19/78/74830572_7185.jpg?1688872459", 
              "https://static1.mingle2.com/images/users/33/41/74830572_339.jpg?1688872459",
               "https://static1.mingle2.com/images/users/12/15/74833197_2539.jpg?1688872318", 
               "https://static1.mingle2.com/images/users/48/65/74833197_2758.jpg?1688872501", 
               "https://static1.mingle2.com/images/users/7/9/74833197_5819.jpg?1688872501", 
               "https://static1.mingle2.com/images/users/91/60/74919929_2816.jpg?1689292634", 
               "https://cdn-v3.justsayhi.com/u/profile_photo/Qrrsdsl9eQN_large_optimized.jpeg", 
               "https://cdn-v3.justsayhi.com/u/profile_photo/2uxWjJQ9WyW_large_optimized.jpeg",
                "https://cdn-v3.justsayhi.com/u/profile_photo/InlcyxUEcYL_large_optimized.jpeg", 
                "https://cdn-v3.justsayhi.com/u/profile_photo/Y13L5bW7mAR_large_optimized.jpeg", 
                "https://static1.mingle2.com/images/users/98/65/74986085_4378.jpg?1689569696", 
                "https://static1.mingle2.com/images/users/41/84/74986085_1317.jpg?1689569861", 
                "https://static1.mingle2.com/images/users/22/8/74986085_5260.jpg?1689569861", 
                "https://cdn-v3.justsayhi.com/u/profile_photo/uKn2_lKJwPI_large_optimized.jpeg",
                 "https://cdn-v3.justsayhi.com/u/profile_photo/AJcPXoMAdIv_large_optimized.jpeg",
                  "https://cdn-v3.justsayhi.com/u/profile_photo/qjky3UK5Q9s_large_optimized.jpeg", 
                  "https://cdn-v3.justsayhi.com/u/profile_photo/v050l__AypJ_large_optimized.jpeg", 
                  "https://static1.mingle2.com/images/users/37/28/75124506_773.jpg?1690197773", 
                  "https://static1.mingle2.com/images/users/89/33/75124506_9591.jpg?1690197840",
                   "https://static1.mingle2.com/images/users/97/40/75124506_2207.jpg?1690197840", 
                   "https://static1.mingle2.com/images/users/70/20/75124506_763.jpg?1690201095", 
                   "https://static1.mingle2.com/images/users/85/24/75124506_953.jpg?1690201095",
                    "https://static1.mingle2.com/images/users/45/20/75214370_226.jpg?1691532209", 
                    "https://static1.mingle2.com/images/users/25/28/75214370_4337.jpg?1691101182", 
                    "https://static1.mingle2.com/images/users/59/59/75214370_1622.jpg?1691101182", 
                    "https://static1.mingle2.com/images/users/97/1/75214370_699.jpg?1691101182", 
                    "https://static1.mingle2.com/images/users/48/50/75380511_1326.jpg?1691360168",
                     "https://static1.mingle2.com/images/users/57/92/75421696_5461.jpg?1691559334", 
                      "https://static1.mingle2.com/images/users/76/49/75421696_3951.jpg?1691559244", 
                      "https://static1.mingle2.com/images/users/62/97/75440625_5131.jpg?1691606627",
                       "https://static1.mingle2.com/images/users/2/42/75464357_3618.jpg?1691708044", 
                       "https://static1.mingle2.com/images/users/1/39/75524581_6119.jpg?1691957300",
                        "https://static1.mingle2.com/images/users/75/5/75524581_9063.jpg?1691957190", 
                        "https://static1.mingle2.com/images/users/32/50/75531357_2855.jpg?1692722108", 
                        "https://static1.mingle2.com/images/users/38/16/75531357_9657.jpg?1692722108", 
                        "https://static1.mingle2.com/images/users/39/51/75589523_4919.jpg?1692227407", 
                        "https://static1.mingle2.com/images/users/20/33/75589523_4849.jpg?1692228970", 
                        "https://static1.mingle2.com/images/users/74/15/75589523_2538.jpg?1692228970",
                         "https://static1.mingle2.com/images/users/91/43/75607600_7405.jpg?1692300644", 
                         "https://static1.mingle2.com/images/users/62/92/75607914_6785.jpg?1692304871",
                          "https://cdn-v3.justsayhi.com/u/profile_photo/n5zaaS_9Wdd_large_optimized.jpeg",
                          
                           "https://cdn-v3.justsayhi.com/u/profile_photo/ISMB814Xgnp_large_optimized.jpeg",
                            "https://static1.mingle2.com/images/users/88/69/75690702_4418.jpg?1692647433",
                             "https://cdn-v3.justsayhi.com/u/profile_photo/nu_Bnlo_Zz6_large_optimized.jpeg", 
                             "https://cdn-v3.justsayhi.com/u/profile_photo/93E1M3ShHxy_large_optimized.jpeg",
                              "https://cdn-v3.justsayhi.com/u/profile_photo/MF-ajvGP6HE_large_optimized.jpeg", 
                              "https://cdn-v3.justsayhi.com/u/profile_photo/NNDwB6sqJYd_large_optimized.jpeg",
                               "https://cdn-v3.justsayhi.com/u/profile_photo/-pT6SNEhLKQ_large_optimized.jpeg", 
                               "https://cdn-v3.justsayhi.com/u/profile_photo/QVwulBhYnbq_large_optimized.jpeg", 
                               "https://cdn-v3.justsayhi.com/u/profile_photo/Ch4rxbilWSB_large_optimized.jpeg",
                                "https://cdn-v3.justsayhi.com/u/profile_photo/p4jnjUvURu-_large_optimized.jpeg", 
                                "https://static1.mingle2.com/images/users/15/8/75898174_7270.jpg?1693579439", 
                                "https://static1.mingle2.com/images/users/52/97/75898174_5824.jpg?1695042202",
                                 "https://static1.mingle2.com/images/users/39/86/75898174_446.jpg?1695042202", 
                                 "https://static1.mingle2.com/images/users/62/17/75957029_8814.jpg?1693878747", 
                                 "https://static1.mingle2.com/images/users/81/19/75957029_2552.jpg?1693877582", 
                                 "https://static1.mingle2.com/images/users/42/46/75957029_3725.jpg?1693877582", 
                                 "https://static1.mingle2.com/images/users/40/92/75957029_8177.jpg?1693877582", 
                                 "https://static1.mingle2.com/images/users/70/6/75957029_7708.jpg?1693877582", 
                                 "https://cdn-v3.justsayhi.com/u/profile_photo/ttT8oMvMVbs_large_optimized.jpeg", 
                                 "https://cdn-v3.justsayhi.com/u/profile_photo/AtZydnusIKG_large_optimized.jpeg", 
                                 "https://cdn-v3.justsayhi.com/u/profile_photo/zyOTm8ganT2_large_optimized.png", 
                                 "https://cdn-v3.justsayhi.com/u/profile_photo/TqdudF0EG22_large_optimized.png", 
                                 "https://static1.mingle2.com/images/users/5/89/76073624_1607.jpg?1694514247", 
                                 "https://static1.mingle2.com/images/users/10/69/76073624_3271.jpg?1694514347",
                                  "https://static1.mingle2.com/images/users/34/34/76073624_4423.jpg?1694514347", 
                                  "https://static1.mingle2.com/images/users/65/29/76083752_8340.jpg?1694515287", 
                                  "https://static1.mingle2.com/images/users/15/5/76083752_9561.jpg?1695042051", 
                                  "https://static1.mingle2.com/images/users/57/8/76083752_8678.jpg?1695042051",
                                  "https://static1.mingle2.com/images/users/21/20/76083752_5413.jpg?1695042051", 
                                  "https://static1.mingle2.com/images/users/86/1/76118382_7589.jpg?1694675829", 
                                  "https://static1.mingle2.com/images/users/35/7/76118382_1140.jpg?1694676033", 
                                  "https://static1.mingle2.com/images/users/42/7/76118382_4970.jpg?1694676033",
                                   "https://static1.mingle2.com/images/users/35/87/76118382_2398.jpg?1694676033", 
                                   "https://static1.mingle2.com/images/users/3/9/76124473_3814.jpg?1694707499", 
                                   "https://static1.mingle2.com/images/users/2/89/76124473_444.jpg?1694707716",
                                    "https://static1.mingle2.com/images/users/85/44/76124473_545.jpg?1694707716",
                                     "https://static1.mingle2.com/images/users/8/22/76124473_6327.jpg?1694707716",
                                      "https://static1.mingle2.com/images/users/81/74/76124473_8832.jpg?1694707716",
                                       "https://static1.mingle2.com/images/users/98/8/76124473_8511.jpg?1694707716",
                                       "https://static1.mingle2.com/images/users/17/94/76129915_8874.jpg?1694750320", 
                                       "https://static1.mingle2.com/images/users/15/63/76142806_4999.jpg?1695581897",
                                        "https://static1.mingle2.com/images/users/15/35/76174204_524.jpg?1694972253", 
                                        "https://static1.mingle2.com/images/users/77/14/76250078_8632.jpg?1695400156",
                                         "https://static1.mingle2.com/images/users/38/4/76250078_9557.jpg?1695402713", 
                                         "https://static1.mingle2.com/images/users/93/95/76250078_2287.jpg?1695402713",
                                          "https://static1.mingle2.com/images/users/84/16/76250078_9938.jpg?1695402713",
                                           "https://cdn-v3.justsayhi.com/u/profile_photo/ZJOf5M8BtvA_large_optimized.jpeg", 
                                           "https://cdn-v3.justsayhi.com/u/profile_photo/fcb9ErpJs56_large_optimized.jpeg",
                                            "https://cdn-v3.justsayhi.com/u/profile_photo/Ayr1Ds_QTsq_large_optimized.jpeg",
                                             "https://cdn-v3.justsayhi.com/u/profile_photo/nbwmCYau-XH_large_optimized.jpeg", 
                                             "https://cdn-v3.justsayhi.com/u/profile_photo/4yxTjaAbR-F_large_optimized.jpeg",
                                              "https://cdn-v3.justsayhi.com/u/profile_photo/8ls9H8Bif3Z_large_optimized.jpeg", 
                                              "https://cdn-v3.justsayhi.com/u/profile_photo/zExqNQUvJcC_large_optimized.jpeg", 
                                              "https://cdn-v3.justsayhi.com/u/profile_photo/0rMNQCfdFvi_large_optimized.jpeg",
                                               "https://cdn-v3.justsayhi.com/u/profile_photo/U1kgzvzEsqo_large_optimized.jpeg",
                                                "https://cdn-v3.justsayhi.com/u/profile_photo/lHajWwNHOtS_large_optimized.jpeg", 
                                                "https://cdn-v3.justsayhi.com/u/profile_photo/v7SKt9tSB8G_large_optimized.jpeg", 
                                                "https://cdn-v3.justsayhi.com/u/profile_photo/un-nsjJeLrr_large_optimized.jpeg", 
                                                "https://cdn-v3.justsayhi.com/u/profile_photo/O-J6IL1K6yT_large_optimized.jpeg", 
                                                "https://cdn-v3.justsayhi.com/u/profile_photo/eoqKYEd1cAR_large_optimized.jpeg", 
                                                "https://cdn-v3.justsayhi.com/u/profile_photo/wT4W_MHKCz5_large_optimized.jpeg", 
                                                "https://cdn-v3.justsayhi.com/u/profile_photo/7HxTmZi3-AJ_large_optimized.jpeg",
                                                 "https://cdn-v3.justsayhi.com/u/profile_photo/rVSILZm5VP8_large_optimized.jpeg", 
                                                 "https://cdn-v3.justsayhi.com/u/profile_photo/Pjfi3PC9W6H_large_optimized.jpeg",
                                                  "https://cdn-v3.justsayhi.com/u/profile_photo/YPoHjyVMD13_large_optimized.jpeg", 
                                                  "https://cdn-v3.justsayhi.com/u/profile_photo/vCyufhyZoUK_large_optimized.jpeg", 
                                                  "https://cdn-v3.justsayhi.com/u/profile_photo/4H06Cwfub-e_large_optimized.jpeg",
                                                   "https://cdn-v3.justsayhi.com/u/profile_video/m1L-dABBdnm_optimized.jpg", 
                                                   "https://cdn-v3.justsayhi.com/u/profile_photo/jIISY39JOUG_large_optimized.jpeg", 
                                                   "https://cdn-v3.justsayhi.com/u/profile_photo/3FFk5m38OsE_large_optimized.jpeg",
                                                    "https://static1.mingle2.com/images/users/5/7/76359426_7401.jpg?1696013555", 
                                                    "https://cdn-v3.justsayhi.com/u/profile_photo/AWyFEbnfJKM_large_optimized.jpeg", 
                                                    "https://cdn-v3.justsayhi.com/u/profile_photo/5_T1LnaeV-S_large_optimized.jpeg"
                                                     "https://cdn-v3.justsayhi.com/u/profile_photo/LJdvspCZI0Y_large_optimized.jpeg", 
                                                     "https://static1.mingle2.com/images/users/46/35/76405621_9539.jpg?1696271904",
                                                      "https://static1.mingle2.com/images/users/91/10/76405621_1363.jpg?1696272954", 
                                                      "https://static1.mingle2.com/images/users/8/37/76405621_741.jpg?1696272954", 
                                                      "https://static1.mingle2.com/images/users/84/66/76405621_9086.jpg?1696272954", 
                                                      "https://static1.mingle2.com/images/users/77/96/76405621_5219.jpg?1696272954", 
                                                      "https://static1.mingle2.com/images/users/78/48/76405621_2650.jpg?1696272954",
                                                       "https://cdn-v3.justsayhi.com/u/profile_photo/ULEl-Dd-obt_large_optimized.jpeg", 
                                                       "https://cdn-v3.justsayhi.com/u/profile_photo/ZjIBuKfIVYn_large_optimized.jpeg",
                                                        "https://cdn-v3.justsayhi.com/u/profile_photo/vdVwDqNxljH_large_optimized.jpeg", 
                                                        "https://static1.mingle2.com/images/users/71/5/76418951_4296.jpg?1696347024", 
                                                        "https://static1.mingle2.com/images/users/51/56/76443680_9784.jpg?1696486938",
                                                         "https://static1.mingle2.com/images/users/99/94/76443680_8859.jpg?1696489502", 
                                                         "https://static1.mingle2.com/images/users/46/32/76443680_5278.jpg?1696489502",
                                                          "https://static1.mingle2.com/images/users/44/72/76443680_9800.jpg?1696489502", 
                                                          "https://static1.mingle2.com/images/users/26/7/76450493_1735.jpg?1696526743",
                                                           "https://static1.mingle2.com/images/users/80/96/76482078_4245.jpg?1697924573", 
                                                           "https://static1.mingle2.com/images/users/31/59/76562295_4071.jpg?1697115431",
                                                            "https://static1.mingle2.com/images/users/49/35/76562295_357.jpg?1697115824",
                                                             "https://static1.mingle2.com/images/users/89/93/76562295_6423.jpg?1697115824", 
                                                             "https://static1.mingle2.com/images/users/68/33/76564108_5231.jpg?1697124682",
                                                              "https://static1.mingle2.com/images/users/12/77/76566649_2791.jpg?1697135888", 
                                                              "https://static1.mingle2.com/images/users/5/0/76572142_5856.jpg?1697167663", 
                                                              "https://static1.mingle2.com/images/users/97/66/76572142_3138.jpg?1697167010",
                                                               "https://static1.mingle2.com/images/users/12/83/76572142_9706.jpg?1697167010", 
                                                               "https://static1.mingle2.com/images/users/39/50/76580572_8640.jpg?1697214208",
                                                                "https://static1.mingle2.com/images/users/41/96/76597262_1292.jpg?1697292861",
                                                                 "https://static1.mingle2.com/images/users/50/32/76597262_6760.jpg?1697296384",
                                                                  "https://static1.mingle2.com/images/users/76/85/76652873_9199.jpg?1697642406",
                                                                   "https://static1.mingle2.com/images/users/56/56/76652873_1100.jpg?1697567204", 
                                                                   "https://static1.mingle2.com/images/users/50/86/76652873_8211.jpg?1697571014", 
                                                                   "https://static1.mingle2.com/images/users/72/86/76652873_6733.jpg?1697570723", 
                                                                   "https://static1.mingle2.com/images/users/77/55/76652873_30.jpg?1697642406",
                                                                    "https://static1.mingle2.com/images/users/48/47/76652873_1525.jpg?1697642406", 
                                                                    "https://static1.mingle2.com/images/users/53/55/76652873_7876.jpg?1697642406",
                                                                     "https://static1.mingle2.com/images/users/66/40/76653907_766.jpg?1697570872", 
                                                                     "https://static1.mingle2.com/images/users/10/2/76682874_419.jpg?1697682707", 
                                                                     "https://static1.mingle2.com/images/users/19/78/76682874_1649.jpg?1697681443", 
                                                                     "https://static1.mingle2.com/images/users/11/63/76682874_7321.jpg?1697681443",
                                                                      "https://static1.mingle2.com/images/users/48/86/76769095_5203.jpg?1697916700", 
                                                                      "https://static1.mingle2.com/images/users/85/9/76924072_1521.jpg?1698287407", 
                                                                      "https://cdn-v3.justsayhi.com/u/profile_photo/jjOD0jlG5vJ_large_optimized.png",
                                                                       "https://cdn-v3.justsayhi.com/u/profile_photo/X_yuEzyd0b7_large_optimized.png", 
                                                                       "https://cdn-v3.justsayhi.com/u/profile_photo/u8uhHAxgX98_large_optimized.png",
                                                                        "https://cdn-v3.justsayhi.com/u/profile_photo/3CWEICjgc_0_large_optimized.png", 
                                                                        "https://static1.mingle2.com/images/users/96/67/77062153_279.jpg?1698986166", 
                                                                        "https://static1.mingle2.com/images/users/60/55/77062153_1532.jpg?1698985546", 
                                                                        "https://static1.mingle2.com/images/users/29/33/77062153_4975.jpg?1698985546",
                                                                         "https://static1.mingle2.com/images/users/47/66/77072026_5570.jpg?1699039817", 
                                                                         "https://static1.mingle2.com/images/users/95/7/77096071_174.jpg?1699174836", 
                                                                         "https://static1.mingle2.com/images/users/84/89/77108106_4941.jpg?1699232582",
                                                                          "https://static1.mingle2.com/images/users/28/94/77124325_4316.jpg?1699330665", 
                                                                           "https://static1.mingle2.com/images/users/46/13/77255559_7840.jpg?1700024907", 
                                                                           "https://static1.mingle2.com/images/users/85/22/77285246_4906.jpg?1700184176", 
                                                                           "https://static1.mingle2.com/images/users/74/59/77296161_588.jpg?1700243968", 
                                                                           "https://static1.mingle2.com/images/users/60/99/77336239_3235.jpg?1700440779", 
                                                                           "https://static1.mingle2.com/images/users/46/40/77352910_7799.jpg?1700531662",
                                                                            "https://static1.mingle2.com/images/users/97/80/77370714_9630.jpg?1700629370",
                                                                             "https://static1.mingle2.com/images/users/54/20/77386489_1681.jpg?1700716218",
                                                                              "https://static1.mingle2.com/images/users/11/3/77386740_6627.jpg?1700717454", 
                                                                              "https://static1.mingle2.com/images/users/68/40/77490601_6086.jpg?1701289040", 
                                                                              "https://static1.mingle2.com/images/users/74/0/77539468_882.jpg?1701570212",
                                                                               "https://static1.mingle2.com/images/users/92/62/77700365_9000.jpg?1702483430", 
                                                                               "https://static1.mingle2.com/images/users/23/15/77700491_9454.jpg?1702486989",
                                                                                "https://static1.mingle2.com/images/users/87/60/77792461_7583.jpg?1703006972", 
                                                                                "https://static1.mingle2.com/images/users/19/29/77813660_8787.jpg?1703119323", 
                                                                                "https://static1.mingle2.com/images/users/35/71/77951855_3858.jpg?1703901790",
                                                                                 "https://static1.mingle2.com/images/users/30/33/77951855_7631.jpg?1703899857", 
                                                                                 "https://static1.mingle2.com/images/users/22/13/77951855_6250.jpg?1703899857",
                                                                                  "https://static1.mingle2.com/images/users/76/51/78148323_150.jpg?1704855327", 
                                                                                  "https://static1.mingle2.com/images/users/53/66/78148323_5542.jpg?1704855507", 
                                                                                  "https://static1.mingle2.com/images/users/29/44/78226734_9930.jpg?1705259991",
                                                                                   "https://static1.mingle2.com/images/users/60/40/78409164_9586.jpg?1706204216",
                                                                                    "https://static1.mingle2.com/images/users/11/19/78414115_1423.jpg?1706239371",
                                                                                     "https://static1.mingle2.com/images/users/53/29/78451466_7029.jpg?1706422371", 
                                                                                     "https://static1.mingle2.com/images/users/99/82/78515391_8634.jpg?1706753146", 
                                                                                     "https://static1.mingle2.com/images/users/24/26/78538515_5919.jpg?1707045543",
                                                                                      "https://static1.mingle2.com/images/users/2/59/78538515_826.jpg?1707047267",
                                                                                       "https://static1.mingle2.com/images/users/38/91/78575565_1800.jpg?1707078811", 
                                                                                       "https://static1.mingle2.com/images/users/40/89/78575869_313.jpg?1707230676", 
                                                                                       "https://static1.mingle2.com/images/users/76/81/78575869_9139.jpg?1707230638", 
                                                                                       "https://static1.mingle2.com/images/users/17/38/78615527_3015.jpg?1707282298", 
                                                                                       "https://static1.mingle2.com/images/users/20/59/78709107_3345.jpg?1707787843",
                                                                                        "https://cdn-v3.justsayhi.com/u/profile_photo/2ZDqbEpdfU9_large_optimized.png", 
                                                                                        "https://cdn-v3.justsayhi.com/u/profile_photo/DqzoIgjzUEi_large_optimized.jpeg", 
                                                                                        "https://static1.mingle2.com/images/users/49/48/78836721_7521.jpg?1708452470",
                                                                                         "https://static1.mingle2.com/images/users/23/73/78838822_525.jpg?1708460996",
                                                                                          "https://static1.mingle2.com/images/users/28/67/78838822_9034.jpg?1708465793",
                                                                                           "https://static1.mingle2.com/images/users/47/62/78838822_4287.jpg?1708465793",
                                                                                            "https://static1.mingle2.com/images/users/60/51/78886755_65.jpg?1708710654", 
                                                                                            "https://static1.mingle2.com/images/users/97/8/78926478_7687.jpg?1708914981",
                                                                                             "https://static1.mingle2.com/images/users/12/99/79052722_9971.jpg?1709744295",
                                                                                              "https://static1.mingle2.com/images/users/95/32/79097316_3007.jpg?1709831146", 
                                                                                              "https://static1.mingle2.com/images/users/83/35/79170187_7805.jpg?1710218228", 
                                                                                              "https://cdn-v3.justsayhi.com/u/profile_photo/9Y3SNMPMX5T_large_optimized.jpeg", 
                                                                                              "https://cdn-v3.justsayhi.com/u/profile_photo/xOlJclhBJjY_large_optimized.jpeg",
                                                                                               "https://cdn-v3.justsayhi.com/u/profile_photo/ZmDaENEj2YZ_large_optimized.jpeg", 
                                                                                               "https://cdn-v3.justsayhi.com/u/profile_photo/MXYjwreVN9v_large_optimized.jpeg", 
                                                                                               "https://cdn-v3.justsayhi.com/u/profile_photo/5Fo-Iv5sX8n_large_optimized.jpeg",
                                                                                                "https://cdn-v3.justsayhi.com/u/profile_photo/u5vTzPvIk3A_large_optimized.jpeg", 
                                                                                                "https://cdn-v3.justsayhi.com/u/profile_photo/mc2XscyE3Cj_large_optimized.jpeg",
                                                                                                 "https://static1.mingle2.com/images/users/9/40/79333732_4694.jpg?1711191192",
                                                                                                  "https://static1.mingle2.com/images/users/82/51/79333732_9157.jpg?1711195093",
                                                                                                   "https://static1.mingle2.com/images/users/48/43/79441627_4677.jpg?1711585939", 
                                                                                                   "https://static1.mingle2.com/images/users/74/40/79442708_4438.jpg?1711589842", 
                                                                                                   "https://static1.mingle2.com/images/users/19/79/79509467_2415.jpg?1711938068",
                                                                                                    "https://cdn-v3.justsayhi.com/u/profile_photo/ZovTyWbIylf_large_optimized.jpeg",
                                                                                                     "https://cdn-v3.justsayhi.com/u/profile_photo/GJEkMMu0XIq_large_optimized.jpeg", 
                                                                                                     "https://cdn-v3.justsayhi.com/u/profile_photo/KQstq9SlMzm_large_optimized.jpeg",
                                                                                                      "https://cdn-v3.justsayhi.com/u/profile_photo/taZggo1VHBx_large_optimized.jpeg", 
                                                                                                      "https://static1.mingle2.com/images/users/17/14/79543052_7332.jpg?1712118638", 
                                                                                                      "https://static1.mingle2.com/images/users/43/2/79562729_6691.jpg?1712160135", 
                                                                                                      "https://static1.mingle2.com/images/users/18/38/79587697_9715.jpg?1712277920", 
                                                                                                      "https://static1.mingle2.com/images/users/93/44/79602161_7875.jpg?1712349051", 
                                                                                                      "https://static1.mingle2.com/images/users/12/81/79609663_3201.jpg?1712385341", 
                                                                                                      "https://static1.mingle2.com/images/users/38/49/79609663_8177.jpg?1712385868", 
                                                                                                      "https://static1.mingle2.com/images/users/6/97/79644548_2561.jpg?1712543316", 
                                                                                                      "https://static1.mingle2.com/images/users/36/75/79654025_5214.jpg?1712594024", 
                                                                                                      "https://static1.mingle2.com/images/users/96/14/79665607_5770.jpg?1712946940",
                                                                                                       "https://cdn-v3.justsayhi.com/u/profile_photo/9jlzdeC4mpN_large_optimized.jpeg", 
                                                                                                       "https://static1.mingle2.com/images/users/42/20/79780037_199.jpg?1713141460",
                                                                                                        "https://static1.mingle2.com/images/users/84/30/79781089_6453.jpg?1713146502",
                                                                                                         "https://static1.mingle2.com/images/users/44/75/79791344_6014.jpg?1713195491",
                                                                                                          "https://static1.mingle2.com/images/users/92/11/79802761_4344.jpg?1713245013",
                                                                                                           "https://static1.mingle2.com/images/users/78/50/79807677_2256.jpg?1713268330", 
                                                                                                           "https://static1.mingle2.com/images/users/4/46/79820979_4476.jpg?1713325195",
                                                                                                            "https://static1.mingle2.com/images/users/39/17/79820979_216.jpg?1713325458",
                                                                                                             "https://static1.mingle2.com/images/users/43/20/79831773_1165.jpg?1713373446",
                                                                                                              "https://static1.mingle2.com/images/users/96/58/79834717_1496.jpg?1713401854", 
                                                                                                              "https://static1.mingle2.com/images/users/99/53/79871140_3690.jpg?1713554740",
                                                                                                               "https://static1.mingle2.com/images/users/48/76/79871241_3386.jpg?1713554839",
                                                                                                                "https://static1.mingle2.com/images/users/73/57/79871840_5774.jpg?1713557102",
                                                                                                                 "https://static1.mingle2.com/images/users/45/13/79895081_7581.jpg?1713663558",
                                                                                                                  "https://static1.mingle2.com/images/users/92/58/79906916_4259.jpg?1713717213", 
                                                                                                                  "https://cdn-v3.justsayhi.com/u/profile_photo/tjiV-2nMLAu_large_optimized.png", 
                                                                                                                  "https://static1.mingle2.com/images/users/62/96/79917860_5344.jpg?1713913720",
                                                                                                                   "https://static1.mingle2.com/images/users/97/46/79917860_265.jpg?1713914809", 
                                                                                                                   "https://static1.mingle2.com/images/users/53/93/79934961_7561.jpg?1713842131", 
                                                                                                                   "https://static1.mingle2.com/images/users/31/64/79943698_2030.jpg?1713887310", 
                                                                                                                   "https://static1.mingle2.com/images/users/8/53/79972162_8052.jpg?1714013031",
                                                                                                                    "https://static1.mingle2.com/images/users/20/96/79973457_279.jpg?1714019741", 
                                                                                                                    "https://static1.mingle2.com/images/users/88/65/79973457_3270.jpg?1714022223",
                                                                                                                     "https://static1.mingle2.com/images/users/1/6/79973889_3510.jpg?1714024362",
                                                                                                                      "https://static1.mingle2.com/images/users/50/39/79973889_2151.jpg?1714024570",
                                                                                                                       "https://static1.mingle2.com/images/users/80/8/79973889_7453.jpg?1714024570", 
                                                                                                                       "https://static1.mingle2.com/images/users/27/59/80017636_8832.jpg?1714276815",
                                                                                                                        "https://static1.mingle2.com/images/users/14/5/80020742_2458.jpg?1714237481",
                                                                                                                         "https://static1.mingle2.com/images/users/76/76/80027769_6488.jpg?1714266004",
                                                                                                                          "https://static1.mingle2.com/images/users/97/3/80046008_4936.jpg?1714349162",
                                                                                                                           "https://static1.mingle2.com/images/users/61/54/80048682_879.jpg?1714362725",
                                                                                                                            "https://static1.mingle2.com/images/users/35/52/80048682_783.jpg?1714364704",
                                                                                                                             "https://static1.mingle2.com/images/users/7/32/80048682_688.jpg?1714364704",
                                                                                                                              "https://cdn-v3.justsayhi.com/u/profile_photo/UMuiYAxj_tZ_large_optimized.png",
                                                                                                                               "https://cdn-v3.justsayhi.com/u/profile_photo/SDCJhGQTR0J_large_optimized.png", 
                                                                                                                               "https://static1.mingle2.com/images/users/83/31/80100731_2225.jpg?1714590815", 
                                                                                                                               "https://cdn-v3.justsayhi.com/u/profile_photo/ni8oWsMy4Z1_large_optimized.png",
                                                                                                                                "https://static1.mingle2.com/images/users/21/50/80103713_7151.jpg?1714599239", 
                                                                                                                                "https://cdn-v3.justsayhi.com/u/profile_photo/E6oVQMWU2n5_large_optimized.png",
                                                                                                                                 "https://cdn-v3.justsayhi.com/u/profile_photo/Iule-Zo6tOR_large_optimized.png",
                                                                                                                                  "https://cdn-v3.justsayhi.com/u/profile_photo/3DRFS0-S0Jn_large_optimized.png",
                                                                                                                                   "https://cdn-v3.justsayhi.com/u/profile_photo/jDpmKW7TnNh_large_optimized.png",
                                                                                                                                    "https://cdn-v3.justsayhi.com/u/profile_photo/YFkjSAGMhHk_large_optimized.png", 
                                                                                                                                    "https://cdn-v3.justsayhi.com/u/profile_photo/h3MN1CqLDKO_large_optimized.png", 
                                                                                                                                    "https://cdn-v3.justsayhi.com/u/profile_photo/HdWL3D1Pz8q_large_optimized.png", 
                                                                                                                                    "https://cdn-v3.justsayhi.com/u/profile_photo/VImtqQARhff_large_optimized.png", 
                                                                                                                                    "https://cdn-v3.justsayhi.com/u/profile_photo/sxMy_5Hu72h_large_optimized.png", 
                                                                                                                                    "https://cdn-v3.justsayhi.com/u/profile_photo/Y4naHTvQpHq_large_optimized.jpeg",
                                                                                                                                     "https://static1.mingle2.com/images/users/67/48/80127405_4174.jpg?1714702771", 
                                                                                                                                     "https://static1.mingle2.com/images/users/90/35/80127758_6076.jpg?1714705109", 
                                                                                                                                     "https://static1.mingle2.com/images/users/60/19/80130134_1851.jpg?1714752330", 
                                                                                                                                     "https://static1.mingle2.com/images/users/15/9/80159737_3398.jpg?1714847616",
                                                                                                                                      "https://cdn-v3.justsayhi.com/u/profile_photo/18rXx2zrNxo_large_optimized.jpeg",
                                                                                                                                       "https://static1.mingle2.com/images/users/40/94/80176306_7930.jpg?1715036923",
                                                                                                                                        "https://static1.mingle2.com/images/users/63/99/80176306_1797.jpg?1714919116",
                                                                                                                                         "https://static1.mingle2.com/images/users/3/39/80176306_1584.jpg?1714919116",
                                                                                                                                          "https://static1.mingle2.com/images/users/90/7/80191104_7751.jpg?1715054210",
                                                                                                                                           "https://static1.mingle2.com/images/users/3/39/80205083_8955.jpg?1715061625", 
                                                                                                                                           "https://static1.mingle2.com/images/users/41/56/80205083_4687.jpg?1715061905",
                                                                                                                                            "https://static1.mingle2.com/images/users/50/30/80205083_857.jpg?1715061905",
                                                                                                                                             "https://static1.mingle2.com/images/users/27/91/80205572_7725.jpg?1715063486",
                                                                                                                                              "https://static1.mingle2.com/images/users/97/89/80205572_5646.jpg?1715063635",
                                                                                                                                               "https://static1.mingle2.com/images/users/76/80/80205572_9503.jpg?1715063635",
                                                                                                                                                "https://cdn-v3.justsayhi.com/u/profile_photo/pf_oGW_W4zj_large_optimized.png",
                                                                                                                                                 "https://cdn-v3.justsayhi.com/u/profile_photo/U-BWibjUFYn_large_optimized.png",
                                                                                                                                                  "https://static1.mingle2.com/images/users/31/60/80222836_5923.jpg?1715139992", 
                                                                                                                                                  "https://cdn-v3.justsayhi.com/u/profile_photo/gMFvoEPeBsJ_large_optimized.png",
                                                                                                                                                   "https://static1.mingle2.com/images/users/96/65/80223968_6223.jpg?1715146582", 
                                                                                                                                                   "https://static1.mingle2.com/images/users/6/39/80223968_367.jpg?1715146700",
                                                                                                                                                    "https://static1.mingle2.com/images/users/52/95/80223968_429.jpg?1715146700",
                                                                                                                                                     "https://static1.mingle2.com/images/users/15/14/80224073_5413.jpg?1715146582",
                                                                                                                                                      "https://static1.mingle2.com/images/users/54/18/80224073_3713.jpg?1715146718",
                                                                                                                                                       "https://static1.mingle2.com/images/users/80/15/80224073_2072.jpg?1715146718",
                                                                                                                                                        "https://static1.mingle2.com/images/users/59/15/80225945_4784.jpg?1715217266",
                                                                                                                                                         "https://static1.mingle2.com/images/users/69/16/80226184_7314.jpg?1715217266",
                                                                                                                                                          "https://static1.mingle2.com/images/users/54/11/80237383_7742.jpg?1715211579",
                                                                                                                                                           "https://static1.mingle2.com/images/users/99/15/80243448_8325.jpg?1715231593",
                                                                                                                                                            "https://static1.mingle2.com/images/users/2/97/80243448_6350.jpg?1715231967",
                                                                                                                                                             "https://static1.mingle2.com/images/users/61/4/80243448_419.jpg?1715231967",
                                                                                                                                                              "https://static1.mingle2.com/images/users/3/32/80244753_7736.jpg?1715235759",
                                                                                                                                                               "https://static1.mingle2.com/images/users/86/61/80244753_2804.jpg?1715237654",
                                                                                                                                                                "https://static1.mingle2.com/images/users/98/68/80244753_7702.jpg?1715237654", 
                                                                                                                                                                "https://static1.mingle2.com/images/users/55/36/80246279_7542.jpg?1715274931", 
                                                                                                                                                                "https://static1.mingle2.com/images/users/85/25/80246279_6300.jpg?1715275160", 
                                                                                                                                                                "https://static1.mingle2.com/images/users/31/66/80249883_873.jpg?1715258897", 
                                                                                                                                                                "https://static1.mingle2.com/images/users/48/48/80273201_590.jpg?1715361740",
                                                                                                                                                                 "https://static1.mingle2.com/images/users/28/92/80283503_1146.jpg?1715406719",
                                                                                                                                                                  "https://static1.mingle2.com/images/users/21/67/80283503_5200.jpg?1715406774", 
                                                                                                                                                                  "https://static1.mingle2.com/images/users/63/76/80304115_1935.jpg?1715519186", 
                                                                                                                                                                  "https://static1.mingle2.com/images/users/75/50/80304265_2292.jpg?1715503519", 
                                                                                                                                                                  "https://static1.mingle2.com/images/users/0/87/80304333_4902.jpg?1715503519",
                                                                                                                                                                   "https://static1.mingle2.com/images/users/92/60/80304880_4308.jpg?1715503894",
                                                                                                                                                                    "https://static1.mingle2.com/images/users/54/12/80304880_9361.jpg?1715505777",
                                                                                                                                                                     "https://static1.mingle2.com/images/users/44/12/80304880_2094.jpg?1715505777",
                                                                                                                                                                      "https://static1.mingle2.com/images/users/25/97/80305751_3617.jpg?1715504490", 
                                                                                                                                                                      "https://static1.mingle2.com/images/users/75/0/80305751_6919.jpg?1715506147",
                                                                                                                                                                       "https://static1.mingle2.com/images/users/7/55/80305751_8269.jpg?1715506147",
                                                                                                                                                                        "https://static1.mingle2.com/images/users/90/78/80305751_5123.jpg?1715506147", 
                                                                                                                                                                        "https://static1.mingle2.com/images/users/3/1/80327752_9212.jpg?1715781402",
                                                                                                                                                                         "https://static1.mingle2.com/images/users/85/25/80327991_37.jpg?1715598715", 
                                                                                                                                                                         "https://static1.mingle2.com/images/users/19/54/80329046_2083.jpg?1715603309",
                                                                                                                                                                          "https://static1.mingle2.com/images/users/63/60/80334707_3779.jpg?1716211550", 
                                                                                                                                                                          "https://static1.mingle2.com/images/users/60/59/80335258_595.jpg?1716209519",
                                                                                                                                                                           "https://static1.mingle2.com/images/users/33/49/80335385_7859.jpg?1716209111", 
                                                                                                                                                                           "https://static1.mingle2.com/images/users/5/71/80335477_4816.jpg?1715717318",
                                                                                                                                                                            "https://cdn-v3.justsayhi.com/u/profile_photo/_pw-7uQKCgf_large_optimized.png",
                                                                                                                                                                             "https://cdn-v3.justsayhi.com/u/profile_photo/uMEfTXqfiat_large_optimized.png",
                                                                                                                                                                              "https://cdn-v3.justsayhi.com/u/profile_photo/i_QUie0VahT_large_optimized.png",
                                                                                                                                                                               "https://cdn-v3.justsayhi.com/u/profile_photo/Rg4kcJzvnCa_large_optimized.png",
                                                                                                                                                                                "https://cdn-v3.justsayhi.com/u/profile_photo/l2dsWdbBZby_large_optimized.png",
                                                                                                                                                                                 "https://cdn-v3.justsayhi.com/u/profile_photo/0IjgCWp9FVz_large_optimized.png",
                                                                                                                                                                                  "https://static1.mingle2.com/images/users/12/87/80340258_6011.jpg?1715647896",
                                                                                                                                                                                   "https://static1.mingle2.com/images/users/8/60/80340499_3929.jpg?1715650996",
                                                                                                                                                                                    "https://static1.mingle2.com/images/users/86/43/80340531_7408.jpg?1715650996",
                                                                                                                                                                                     "https://static1.mingle2.com/images/users/17/83/80342172_9028.jpg?1715656908", 
                                                                                                                                                                                     "https://static1.mingle2.com/images/users/46/61/80342282_7386.jpg?1715656908",
                                                                                                                                                                                      "https://static1.mingle2.com/images/users/70/55/80342282_6781.jpg?1715657550",
                                                                                                                                                                                       "https://static1.mingle2.com/images/users/82/87/80342282_6198.jpg?1715657550", 
                                                                                                                                                                                       "https://static1.mingle2.com/images/users/19/2/80342308_5761.jpg?1715657179",
                                                                                                                                                                                        "https://static1.mingle2.com/images/users/78/89/80342756_9995.jpg?1715659744",
                                                                                                                                                                                         "https://static1.mingle2.com/images/users/3/97/80343459_6187.jpg?1719224291", 
                                                                                                                                                                                         "https://static1.mingle2.com/images/users/18/39/80343459_5084.jpg?1719224192", 
                                                                                                                                                                                         "https://static1.mingle2.com/images/users/76/14/80343459_3216.jpg?1719224192",
                                                                                                                                                                                          "https://static1.mingle2.com/images/users/48/44/80358403_9866.jpg?1715724479",
                                                                                                                                                                                           "https://static1.mingle2.com/images/users/91/6/80359392_8615.jpg?1715727796",
                                                                                                                                                                                            "https://static1.mingle2.com/images/users/9/62/80361572_9165.jpg?1715737618", 
                                                                                                                                                                                            "https://static1.mingle2.com/images/users/60/82/80362590_8121.jpg?1715740918",
                                                                                                                                                                                             "https://static1.mingle2.com/images/users/5/80/80362944_205.jpg?1715743970",
                                                                                                                                                                                              "https://static1.mingle2.com/images/users/26/24/80363756_5507.jpg?1715748637",
                                                                                                                                                                                               "https://static1.mingle2.com/images/users/93/90/80364645_6597.jpg?1715753428",
                                                                                                                                                                                                "https://static1.mingle2.com/images/users/71/42/80365021_3395.jpg?1715753483",
                                                                                                                                                                                                 "https://static1.mingle2.com/images/users/28/78/80379919_1546.jpg?1715816170", 
                                                                                                                                                                                                 "https://static1.mingle2.com/images/users/46/25/80379933_837.jpg?1715816170", 
                                                                                                                                                                                                 "https://static1.mingle2.com/images/users/61/59/80381053_2985.jpg?1715821554", 
                                                                                                                                                                                                 "https://static1.mingle2.com/images/users/37/75/80402465_5512.jpg?1715912758",
                                                                                                                                                                                                  "https://static1.mingle2.com/images/users/33/69/80422250_7354.jpg?1716004472",
                                                                                                                                                                                                   "https://static1.mingle2.com/images/users/52/18/80423152_2610.jpg?1716004150",
                                                                                                                                                                                                    "https://static1.mingle2.com/images/users/53/47/80424810_2787.jpg?1716052764", 
                                                                                                                                                                                                    "https://static1.mingle2.com/images/users/38/31/80424975_1166.jpg?1716021493",
                                                                                                                                                                                                     "https://static1.mingle2.com/images/users/6/58/80432432_4372.jpg?1716048197",
                                                                                                                                                                                                      "https://static1.mingle2.com/images/users/2/41/80432432_1372.jpg?1716046473",
                                                                                                                                                                                                       "https://static1.mingle2.com/images/users/1/63/80432432_5492.jpg?1716047078", 
                                                                                                                                                                                                       "https://static1.mingle2.com/images/users/47/15/80432432_5201.jpg?1716047242", 
                                                                                                                                                                                                       "https://static1.mingle2.com/images/users/6/35/80433559_1150.jpg?1716224406",
                                                                                                                                                                                                        "https://static1.mingle2.com/images/users/92/0/80433559_6724.jpg?1716224226",
                                                                                                                                                                                                         "https://static1.mingle2.com/images/users/3/21/80441457_6614.jpg?1716081451",
                                                                                                                                                                                                          "https://static1.mingle2.com/images/users/96/19/80442780_4884.jpg?1716086285",
                                                                                                                                                                                                           "https://static1.mingle2.com/images/users/90/45/80477209_9352.jpg?1716230221",
                                                                                                                                                                                                            "https://static1.mingle2.com/images/users/10/3/80491544_9911.jpg?1716293962",
                                                                                          "https://static1.mingle2.com/images/users/13/45/80491544_5742.jpg?1716294126", 
                                                                                          "https://static1.mingle2.com/images/users/50/13/80505811_8854.jpg?1716353003", 
                                                                                          "https://static1.mingle2.com/images/users/5/31/80524402_8200.jpg?1716429931", 
                                                                                          "https://static1.mingle2.com/images/users/47/10/80545963_9655.jpg?1716520035", 
                                                                                          "https://static1.mingle2.com/images/users/47/41/80585390_3231.jpg?1716682128", 
                                                                                          "https://static1.mingle2.com/images/users/2/80/80622325_3718.jpg?1716829398", 
                                                                                          "https://static1.mingle2.com/images/users/85/73/80662714_7831.jpg?1717000069", 
                                                                                          "https://static1.mingle2.com/images/users/72/64/80680811_1051.jpg?1719352661", 
                                                                                          "https://static1.mingle2.com/images/users/50/72/80680811_3289.jpg?1717247580", 
                                                                                          "https://static1.mingle2.com/images/users/1/83/80680811_8779.jpg?1717247584", 
                                                                                          "https://static1.mingle2.com/images/users/16/70/80680811_7487.jpg?1718589312", 
                                                                                          "https://cdn-v3.justsayhi.com/u/profile_photo/qZlAIGgHofS_large_optimized.png",
                                                                                           "https://cdn-v3.justsayhi.com/u/profile_photo/PYEfwVKERul_large_optimized.png", 
                                                                                           "https://cdn-v3.justsayhi.com/u/profile_photo/JalEOW3-xS__large_optimized.png", 
                                                                                           "https://static1.mingle2.com/images/users/33/12/80765446_1392.jpg?1717433613", 
                                                                                           "https://static1.mingle2.com/images/users/22/89/80790112_8410.jpg?1717538367", 
                                                                                           "https://static1.mingle2.com/images/users/11/98/80790112_6044.jpg?1717538345", 
                                                                                           "https://static1.mingle2.com/images/users/39/31/80792725_8378.jpg?1717547300", 
"https://static1.mingle2.com/images/users/3/56/80792725_7380.jpg?1717547300", 
"https://static1.mingle2.com/images/users/8/94/80792962_2469.jpg?1717547662", 
"https://static1.mingle2.com/images/users/5/28/80845377_4732.jpg?1717778069", 
"https://static1.mingle2.com/images/users/70/98/80910155_1245.jpg?1718043562", 
"https://static1.mingle2.com/images/users/6/98/80961228_1280.jpg?1718265317", 
"https://static1.mingle2.com/images/users/69/51/80961228_9704.jpg?1718616646", 
"https://static1.mingle2.com/images/users/89/99/80961228_3796.jpg?1718616646", 
"https://static1.mingle2.com/images/users/18/97/80961228_7897.jpg?1718616646", 
"https://static1.mingle2.com/images/users/92/10/81041838_5339.jpg?1718670608", 
"https://static1.mingle2.com/images/users/16/75/81042574_6623.jpg?1718629537", 
"https://static1.mingle2.com/images/users/3/37/81526995_8276.jpg?1720628797", 
"https://static1.mingle2.com/images/users/66/55/81526995_7557.jpg?1720629272", 
"https://static1.mingle2.com/images/users/68/14/81526995_9467.jpg?1720629272", 
"https://static1.mingle2.com/images/users/82/41/81971267_9933.jpg?1722413169",
"https://static1.mingle2.com/images/users/35/18/81971328_9830.jpg?1722400439"
                                                                                                                                                                                                               ]


# Function to download an image and return it as a PIL image
def download_image(url):
    response = requests.get(url)
    if response.status_code == 200:
        return PILImage.open(BytesIO(response.content))
    else:
        return None

# Function to analyze image for Gender (replace with actual implementation)
def analyze_gender(image_path):
    # Example request
    response = requests.post(
        'https://nucleuz-dev.swiftsecurity.ai/api/images/detect-gender',
        files={'file': open(image_path, 'rb')}
    )
    if response.status_code == 200:
        return response.json().get("value", "Unknown")
    return "Unknown"

# Function to analyze image for Nudity (replace with actual implementation)
def analyze_nudity(image_path):
    # Example request
    response = requests.post(
        'https://nucleuz-dev.swiftsecurity.ai/api/images/classify',
        files={'file': open(image_path, 'rb')}
    )
    if response.status_code == 200:
        result = response.json()
        return "Yes" if result.get("nudity", False) else "No"
    
    return "Unknown"

# Create a new Excel workbook and add a worksheet
wb = Workbook()
ws = wb.active
ws.title = "Image Analysis"

# Set column headers
ws.append(["URL", "Image", "Filename", "Gender", "Race", "Nudity"])

# Set column widths
ws.column_dimensions['A'].width = 75  # Adjust URL column width
ws.column_dimensions['B'].width = 20  # Adjust image column width
ws.column_dimensions['C'].width = 15  # Adjust Filename column width
ws.column_dimensions['D'].width = 15  # Adjust Gender column width
ws.column_dimensions['E'].width = 15  # Adjust Race column width
ws.column_dimensions['F'].width = 15  # Adjust Nudity column width

# Set a fixed row height for images
row_height = 100

# Create a folder to save images
os.makedirs('images', exist_ok=True)

# Iterate over the URLs
for index, url in enumerate(urls):
    img = download_image(url)
    if img:
        # Save the image locally and add it to the Excel sheet
        img_filename = f"image_{index+1}.png"
        img_path = f"./images/{img_filename}"
        img.save(img_path)

        # Get the image dimensions
        image_width, image_height = img.size

        # Resize image if needed (ensure it fits within the column width)
        max_image_width = 100
        max_image_height = 100
        if image_width > max_image_width or image_height > max_image_height:
            img.thumbnail((max_image_width, max_image_height), PILImage.LANCZOS)
            img_width, img_height = img.size
        else:
            img_width, img_height = image_width, image_height

        # Save the resized image
        img.save(img_path)

        # Load the image into openpyxl
        img_excel = Image(img_path)
        img_excel.width = img_width  # Adjust the size as needed
        img_excel.height = img_height

        # Analyze the image for Gender and Nudity
        gender = analyze_gender(img_path)
        nudity = analyze_nudity(img_path)

        # Append the URL, image, filename, and analysis to the worksheet
        ws[f"A{index+2}"] = url
        ws.add_image(img_excel, f"B{index+2}")
        ws[f"C{index+2}"] = img_filename
        ws[f"D{index+2}"] = gender
        ws[f"E{index+2}"] = "Unknown"  # Placeholder for Race if needed
        ws[f"F{index+2}"] = nudity

        # Set the row height to match the image height
        ws.row_dimensions[index+2].height = img_height  # Add some padding for readability
    else:
        ws.append([url, "Failed to download", "", "", "", ""])

# Save the workbook
wb.save("image_analysis.xlsx")

# To run the code
# pip install requests openpyxl pillow
# python download_image_with_api.py